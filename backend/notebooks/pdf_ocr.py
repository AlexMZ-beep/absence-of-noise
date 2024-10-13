import os
import PyPDF2
import docx
import openpyxl
import easyocr
import ocrmypdf
import tempfile
from google.colab import files
import subprocess


def process_pdf_with_ocr(input_pdf: str, output_pdf: str):
    ocrmypdf.ocr(input_pdf, output_pdf, deskew=True, skip_text=True, language='rus')


def process_pdf(file_path: str) -> str:
    with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as temp_pdf:
        temp_pdf_path = temp_pdf.name

    process_pdf_with_ocr(file_path, temp_pdf_path)

    with open(temp_pdf_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        text = ''
        for page in pdf_reader.pages:
            text += page.extract_text()

    os.remove(temp_pdf_path)
    return text


def process_word(file_path: str) -> str:
    doc = docx.Document(file_path)
    return "\n".join([paragraph.text for paragraph in doc.paragraphs])


def process_excel(file_path: str) -> str:
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    text = ""
    for sheet in workbook.sheetnames:
        text += f"Sheet: {sheet}\n"
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += "\t".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        text += "\n"
    return text


def process_image(file_path: str) -> str:
    reader = easyocr.Reader(['ru'])
    result = reader.readtext(file_path)
    return "\n".join([detection[1] for detection in result])


def detect_file_type(file_path: str) -> str:
    _, extension = os.path.splitext(file_path.lower())
    if extension == '.pdf':
        return 'pdf'
    elif extension in ['.doc', '.docx']:
        return 'word'
    elif extension in ['.xls', '.xlsx']:
        return 'excel'
    elif extension in ['.jpg', '.jpeg', '.png', '.bmp', '.tiff']:
        return 'image'
    else:
        return 'unknown'


def process_file(file_path: str) -> str:
    file_type = detect_file_type(file_path)

    if file_type == 'pdf':
        return process_pdf(file_path)
    elif file_type == 'word':
        return process_word(file_path)
    elif file_type == 'excel':
        return process_excel(file_path)
    elif file_type == 'image':
        return process_image(file_path)
    else:
        return f"Unsupported file type: {file_path}"


def main():
    # file_path = 'CCF_000158.pdf'
    # file_path = 'договор.docx'
    file_path = 'Коллективный договор.pdf'
    output_txt = 'output.txt'

    if not os.path.isfile(file_path):
        print(f"The specified file does not exist: {file_path}")
        print("Please upload the file.")
        uploaded = files.upload()
        if file_path not in uploaded:
            print("The required file was not uploaded.")
            return

    text = process_file(file_path)

    with open(output_txt, 'w', encoding='utf-8') as f:
        f.write(text)

    print(f"Text extraction complete. Content saved to {output_txt}")

    # Display the contents of the output file
    with open(output_txt, 'r', encoding='utf-8') as f:
        print(f.read())


if __name__ == "__main__":
    main()
